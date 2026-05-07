#!/usr/bin/env python3
"""
Debug script to check openpyxl installation and imports
"""

import sys
import os

print("=" * 70)
print("PYTHON DEBUG INFORMATION")
print("=" * 70)
print(f"Python Version: {sys.version}")
print(f"Python Executable: {sys.executable}")
print(f"Python Path: {sys.path}")
print()

# Test 1: Check if openpyxl module can be found
print("=" * 70)
print("TEST 1: Checking if openpyxl is installed...")
print("=" * 70)

try:
    import openpyxl
    print(f"✅ SUCCESS: openpyxl is installed")
    print(f"   Version: {openpyxl.__version__}")
    print(f"   Location: {openpyxl.__file__}")
except ImportError as e:
    print(f"❌ FAILED: openpyxl import error")
    print(f"   Error: {e}")
    sys.exit(1)

print()

# Test 2: Check individual style imports
print("=" * 70)
print("TEST 2: Checking openpyxl.styles imports...")
print("=" * 70)

style_imports = [
    ("Font", "from openpyxl.styles import Font"),
    ("PatternFill", "from openpyxl.styles import PatternFill"),
    ("Alignment", "from openpyxl.styles import Alignment"),
    ("Border", "from openpyxl.styles import Border"),
    ("Side", "from openpyxl.styles import Side"),
    ("get_column_letter", "from openpyxl.utils import get_column_letter"),
]

for name, import_statement in style_imports:
    try:
        exec(import_statement)
        print(f"✅ {name}: OK")
    except Exception as e:
        print(f"❌ {name}: FAILED - {e}")

print()

# Test 3: Full combined import
print("=" * 70)
print("TEST 3: Full combined import (as in your script)...")
print("=" * 70)

EXCEL_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    print(f"✅ openpyxl version: {openpyxl.__version__}")
    print("✅ All imports successful")
    EXCEL_AVAILABLE = True
    print(f"✅ EXCEL_AVAILABLE = {EXCEL_AVAILABLE}")

except ImportError as e:
    print(f"❌ ImportError: {e}")
    EXCEL_AVAILABLE = False
    print(f"❌ EXCEL_AVAILABLE = {EXCEL_AVAILABLE}")
    
except Exception as e:
    print(f"❌ Unexpected Exception: {e}")
    EXCEL_AVAILABLE = False
    print(f"❌ EXCEL_AVAILABLE = {EXCEL_AVAILABLE}")

print()

# Test 4: Create a simple workbook
print("=" * 70)
print("TEST 4: Creating a test workbook...")
print("=" * 70)

if EXCEL_AVAILABLE:
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Test"
        
        # Try to use styles
        ws['A1'].font = Font(bold=True, color="FF0000")
        ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        print("✅ Workbook creation and styling: OK")
        print(f"   Test workbook has {len(wb.sheetnames)} sheet(s)")
        
    except Exception as e:
        print(f"❌ Workbook creation failed: {e}")
else:
    print("⚠️  Skipped (EXCEL_AVAILABLE = False)")

print()
print("=" * 70)
print("SUMMARY")
print("=" * 70)
if EXCEL_AVAILABLE:
    print("✅ openpyxl is properly installed and working!")
else:
    print("❌ openpyxl is NOT working. Run:")
    print("   pip uninstall openpyxl -y")
    print("   pip install openpyxl")
    print("   python debug_imports.py")
